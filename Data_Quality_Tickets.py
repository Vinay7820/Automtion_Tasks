import re
import sys
from openpyxl import Workbook
import openpyxl
from openpyxl import load_workbook
import smtplib
from jira import JIRA
from openpyxl.styles import Alignment
from openpyxl.styles import Color, PatternFill
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from config import config

mydate = datetime.now()
Year = (datetime.now().strftime('%Y'))
Month = (datetime.now().strftime('%m'))
Date = (datetime.now().strftime('%d'))
year_month_format = '{0}-{1}'.format(Year, Month)
start_Date = 1
completed = False

try:
    print("Logging into JIRA")
    jira = JIRA(basic_auth=(config['jira_username'], config['jira_password']), options={'server': 'https://jira.move.com'})
    print("Logged into JIRA succesfully")
except Exception as e:
    print("Not Able to connect to JIRA. Please check your connection and try again!")
    print("Exception - ", e)
    sys.exit()

jra = jira.project('DE')

workbook_name = 'Ticket_Counts.xlsx'
wb = load_workbook(workbook_name)
ws1 = wb['List of tickets']
previous_execution_month = ws1.cell(row = 2, column = 15)

#Function for querying DQ Created issues by category and label in JIRA
def DQ_Created(query):
    None_Leads_Count = 0
    None_Other_Count = 0
    None_Page_Views_Count = 0
    None_UUs_Count = 0

    AWS_Leads_Count = 0
    AWS_Other_Count = 0
    AWS_Page_Views_Count = 0
    AWS_UUs_Count = 0
    EDW_Leads_Count = 0
    EDW_Other_Count = 0
    EDW_Page_Views_Count = 0
    EDW_UUs_Count = 0
    data_category = ""

    key_list = []
    summary_list = []
    status_list = []
    created_date_list = []
    updated_date_list = []
    label_list = []
    assignee_list = []
    reporter_list = []
    data_category_list = []

    for issue in jira.search_issues(query):
        raw_data = issue.raw
        key = issue.key
        key_list.append(issue.key)
        label = issue.fields.labels
        print("LABEL===", label)

        if len(issue.fields.labels) >= 1:
            label_list.append(issue.fields.labels[0])
        elif len(issue.fields.labels) == 0:
            label_list.append("None")

        #####Values are appended to the list because - Sheet 3 holds list of tickets
        status_list.append(issue.fields.status.name)
        created_date_list.append(issue.fields.created)
        updated_date_list.append(issue.fields.updated)
        summary_list.append(issue.fields.summary)
        assignee_list.append(issue.fields.assignee.name)
        reporter_list.append(issue.fields.reporter.name)

        searchObj = re.search(r"('value': 'Page Views'| 'value': 'Leads'| 'value': 'Other'| 'value': 'UUs')", str(raw_data))
        if searchObj:
            remove_quotes = (searchObj.group().replace("'", ""))
            data_category = remove_quotes.split(': ')[1]
            data_category_list.append(data_category)
        else:
            print("There is no Data Category associated with this issue, None will be inserted! Please Check Manually", key)
            data_category_list.append("None")


        if data_category in ["Other", "Page Views", "UUs", "Leads"] and label in [['AWS'], ['EDW'],['edw'],['aws'], ['BDR-ISSUES','AWS'],['BDR-ISSUES','EDW']]:
            if data_category == "Leads" and label in [['AWS'], ['aws'], ['BDR-ISSUES', 'AWS']]:
                AWS_Leads_Count = AWS_Leads_Count + 1
            if data_category == "Leads" and label in [['EDW'], ['edw'], ['BDR-ISSUES', 'EDW']]:
                EDW_Leads_Count = EDW_Leads_Count + 1

            if data_category == "Other" and label in [['AWS'], ['aws'], ['BDR-ISSUES', 'AWS']]:
                AWS_Other_Count = AWS_Other_Count + 1
            if data_category == "Other" and label in [["EDW"], ['edw'], ['BDR-ISSUES', 'EDW']]:
                EDW_Other_Count = EDW_Other_Count + 1

            if data_category == "Page Views" and label in [['AWS'], ['aws'], ['BDR-ISSUES', 'AWS']]:
                AWS_Page_Views_Count = AWS_Page_Views_Count + 1
            if data_category == "Page Views" and label in [["EDW"], ['edw'], ['BDR-ISSUES', 'EDW']]:
                EDW_Page_Views_Count = EDW_Page_Views_Count + 1

            if data_category == "UUs" and label in [['AWS'], ['aws'], ['BDR-ISSUES', 'AWS']]:
                None_UUs_Count = AWS_UUs_Count + 1
            if data_category == "UUs" and label in [["EDW"], ['edw'], ['BDR-ISSUES', 'EDW']]:
                EDW_UUs_Count = EDW_UUs_Count + 1

        # issues with label = None are appended to the sheet 2
        elif data_category in ["Other", "Page Views", "UUs", "Leads"] and label == []:
            if data_category == "Leads":
                None_Leads_Count = None_Leads_Count + 1

            if data_category == "Other":
                None_Other_Count = None_Other_Count + 1

            if data_category == "Page Views":
                None_Page_Views_Count = None_Page_Views_Count + 1

            if data_category == "UUs":
                None_UUs_Count = None_UUs_Count + 1
        else:
            print("There is no Data Category associated for the issue with the key {0}. Please Check!!!".format(key))

    column_name = ['Key', 'Label', 'Summary', 'Data_Category', 'Status', 'Assignee', 'Reporter', 'Created','Updated']
    workbook_name = 'Ticket_Counts.xlsx'
    wb = load_workbook(workbook_name)
    ws1 = wb['List of tickets']

    previous_execution_month = ws1.cell(row = 2, column = 15)
    # print(previous_execution_month.value)

    Key_column = ws1['B']
    excel_key_issue = []
    for x in range(len(Key_column)):
        excel_key_issue.append(Key_column[x].value)

    # print(excel_key_issue)
    # print(key_list)
    # print(label_list)
    # print(summary_list)

    #print("DATA CATEGORY LIST", data_category_list)

    def form_issues(key_lis, key_already_presen, index):
        print("Inside form issues function")
        key_list = [e for e in key_lis if e not in key_already_presen]
        print(key_list)
        #for index in sorted(index_to_remov, reverse=True):
        del label_list[index]
        del status_list[index]
        del summary_list[index]
        del data_category_list[index]
        del created_date_list[index]
        del updated_date_list[index]
        del assignee_list[index]
        del reporter_list[index]

        # print(label_list)
        # print(status_list)
        # print(summary_list)
        # print(data_category_list)
        # print(created_date_list)
        # print(updated_date_list)
        # print(reporter_list)
        # print(assignee_list)

        return key_list, label_list,summary_list,data_category_list, status_list,assignee_list,reporter_list,created_date_list,updated_date_list


    index_to_remove = []
    key_already_present = []

    for key in key_list:
        if key in excel_key_issue:
            print("Removing Duplicate Value")
            print(key)
            index_to_remove = key_list.index(key)
            key_already_present.append(key)
            key_list, label_list, summary_list, data_category_list, status_list, assignee_list, reporter_list, created_date_list, updated_date_list = form_issues(key_list, key_already_present, index_to_remove)
        else:
            pass

    # print("INDEX TO REMOVE ==== ", index_to_remove)
    # print("KEY PRESENT IN EXCEL ====", key_already_present)
    #
    # print(Month)
    # print(previous_execution_month)

    if Month != (previous_execution_month.value):
        max = ws1.max_row
        #print(max)
        key_row = ws1.max_row + 1
        status_row = ws1.max_row + 1
        created_row = ws1.max_row + 1
        updated_row = ws1.max_row + 1
        category_row = ws1.max_row + 1
        label_row = ws1.max_row + 1
        summary_row = ws1.max_row + 1
        assignee_row = ws1.max_row + 1
        reporter_row = ws1.max_row + 1
        start_column = 1
        max = max + 1
        #print(max)

        ws1.merge_cells(start_row=max + 1, start_column=1, end_row=max + 2,
                        end_column=10)  ##Merging the cells according to the excel
        ws1.cell(row=max + 1, column=1).value = 'List of Tickets for {0} month {1}'.format(mydate.strftime("%B"), Year)
        currentCell = ws1.cell(row=max + 1, column=1)
        currentCell.alignment = Alignment(horizontal='center', vertical='center')  ##Aliging the value at the centre
        currentCell.fill = PatternFill(fill_type='solid', start_color='FF6347', end_color='FF6347')  ##styling excel cells
        ws1.cell(row=max + 4, column=start_column).value = '{0}-{1}'.format(Year, Month)

        for header in column_name:
            ws1.cell(row=max + 3, column=start_column + 1).value = header
            start_column += 1

        start_column = 1

        for key in key_list:
            ws1.cell(row=key_row+4, column=start_column+1).value = key
            key_row += 1

        for label in label_list:
            ws1.cell(row=label_row+4, column=start_column + 2).value = label
            label_row += 1

        for summary in summary_list:
            ws1.cell(row=summary_row+4, column=start_column + 3).value = summary
            summary_row += 1

        for category in data_category_list:
            ws1.cell(row=category_row + 4, column=start_column + 4).value = category
            category_row += 1

        for status in status_list:
            ws1.cell(row=status_row + 4, column=start_column + 5).value = status
            status_row += 1

        for assignee in assignee_list:
            ws1.cell(row=assignee_row + 4, column=start_column + 6).value = assignee
            assignee_row += 1

        for reporter in reporter_list:
            ws1.cell(row=reporter_row + 4, column=start_column + 7).value = reporter
            reporter_row += 1

        for created in created_date_list:
            ws1.cell(row=created_row + 4, column=start_column + 8).value = created
            created_row += 1

        for updated in updated_date_list:
            ws1.cell(row=updated_row + 4, column=start_column + 9).value = updated
            updated_row += 1

        value = ['xxxxxxx','xxxxxxxx','xxxxxxxx','xxxxxxxx','##### End of Month #####', mydate.strftime("%B"), year_month_format, 'xxxxxxx','xxxxxxxx','xxxxxxxx','xxxxxxx']

        max = ws1.max_row
        for i in range(1, len(value)):
            ws1.cell(row=max + 1,column=i).value = value[i]
            currentCell = ws1.cell(row=max + 1, column=i)  # or currentCell = ws['A1']
            currentCell.alignment = Alignment(horizontal='center', vertical='center')
            currentCell.fill = PatternFill(fill_type='solid', start_color='FF6347', end_color='FF6347')

            currentCell = ws1.cell(row=max + 2, column=i)  # or currentCell = ws['A1']
            currentCell.fill = PatternFill(fill_type='solid', start_color='000000', end_color='000000')

        wb.save(filename=workbook_name)
        print("Completed Writing List of tickets  Issues to the Excel File")

    else:

        workbook_name = 'Ticket_Counts.xlsx'
        wb = load_workbook(workbook_name)
        ws1 = wb['List of tickets']


        #print("Entered Else Part")
        # print(key_list)
        # print(label_list)
        # print(summary_list)
        # print(data_category_list)
        # print(status_list)
        # print(assignee_list)
        # print(reporter_list)
        # print(created_date_list)
        # print(updated_date_list)

        max = ws1.max_row
        #print(max)
        line_to_start_inserting = max - 1
        print("Line_to_Start_inserting", line_to_start_inserting)
        print("length of key list", len(key_list))
        ws1.insert_rows(line_to_start_inserting, len(key_list))
        print('Inserted {0} rows'.format(len(key_list)) )

        key_row = line_to_start_inserting
        status_row = line_to_start_inserting
        created_row =line_to_start_inserting
        updated_row = line_to_start_inserting
        category_row = line_to_start_inserting
        label_row = line_to_start_inserting
        summary_row = line_to_start_inserting
        assignee_row = line_to_start_inserting
        reporter_row = line_to_start_inserting
        start_column = 1

        for key in key_list:
            ws1.cell(row=key_row, column=start_column+1).value = key
            key_row += 1

        for label in label_list:
            ws1.cell(row=label_row, column=start_column + 2).value = label
            label_row += 1

        for summary in summary_list:
            ws1.cell(row=summary_row, column=start_column + 3).value = summary
            summary_row += 1

        for category in data_category_list:
            ws1.cell(row=category_row, column=start_column + 4).value = category
            category_row += 1

        for status in status_list:
            ws1.cell(row=status_row, column=start_column + 5).value = status
            status_row += 1

        for assignee in assignee_list:
            ws1.cell(row=assignee_row, column=start_column + 6).value = assignee
            assignee_row += 1

        for reporter in reporter_list:
            ws1.cell(row=reporter_row, column=start_column + 7).value = reporter
            reporter_row += 1

        for created in created_date_list:
            ws1.cell(row=created_row, column=start_column + 8).value = created
            created_row += 1

        for updated in updated_date_list:
            ws1.cell(row=updated_row, column=start_column + 9).value = updated
            updated_row += 1

        wb.save(filename=workbook_name)
        #print("Else Part")

    AWS_Subtotal = AWS_Leads_Count + AWS_UUs_Count + AWS_Page_Views_Count + AWS_Other_Count
    EDW_Subtotal = EDW_UUs_Count + EDW_Page_Views_Count + EDW_Other_Count + EDW_Leads_Count
    Grandtotal = AWS_Subtotal + EDW_Subtotal
    Grandtotal_none_count = None_Leads_Count + None_Other_Count + None_Page_Views_Count + None_UUs_Count
    return [EDW_Leads_Count,EDW_Other_Count, EDW_Page_Views_Count,EDW_UUs_Count, EDW_Subtotal,AWS_Leads_Count,AWS_Other_Count,AWS_Page_Views_Count,AWS_UUs_Count,AWS_Subtotal,Grandtotal, None_Leads_Count, None_Other_Count, None_Page_Views_Count, None_UUs_Count, Grandtotal_none_count]

print("Querying the issues for DQ Created")
query = 'project = DE AND issuetype = "Data Quality" AND status in (Blocked, "To Do", Deferred, "Waiting on Others", "In Progress", Open, "In Review", Reopened, Resolved, Closed, Done, "In Review / Validation", "Waiting for Others", "User Misunderstanding", Fixed) AND created >= {0}-{1}-{2} AND created <= {3}-{4}-{5} ORDER BY key ASC, status ASC, cf[10500] ASC, created DESC, due DESC'.format(Year,Month,start_Date,Year,Month,Date)
print("Query = ", query)
DQ_created_result = DQ_Created(query)
Total = DQ_created_result[10]
total_none = DQ_created_result[-1]

#Function for querying DQ open issue, only Closed and open Tickets count are analyzed here.
def DQ_Open(query):
    global Total
    closed_tickets = 0
    open_tickets = 0
    for issue in jira.search_issues(query):
        status = issue.fields.status.name

        if status == "Done":
            closed_tickets = closed_tickets + 1
        else:
            open_tickets = open_tickets + 1

    closed_tickets = Total - open_tickets

    return [open_tickets, closed_tickets]

print("Querying for the issues for DQ Open")
query = 'project in (DE, BISL) AND issuetype = "Data Quality" AND status in (Blocked, "To Do", Deferred, "Waiting on Others", "In Progress", Open, Reopened, "In Review / Validation", "Waiting for Others") AND created >= {0}-{1}-{2} AND created <= {3}-{4}-{5} ORDER BY created DESC, key ASC, status ASC, cf[10500] ASC, due DESC'.format(Year, Month, start_Date, Year, Month, Date)
print("Query = ", query)
DQ_open_result= DQ_Open(query)

##Framing the resultant list according to the cells in the excel, and writing pattern.
DQ_Result = []
DQ_created_result_EDW = DQ_created_result[:5]   ###EDW Subtotal
DQ_created_result_EDW.append(DQ_created_result[10])  ###Grand Total
DQ_created_result_EDW.extend(DQ_open_result)    ####Open and Close
DQ_created_result_AWS = DQ_created_result[5:10]
DQ_Result.append(DQ_created_result_EDW)
DQ_Result.append(DQ_created_result_AWS)
DQ_Null_Report = DQ_created_result[11:]

workbook_name = 'Ticket_Counts.xlsx'
wb = load_workbook(workbook_name)
ws1 = wb['By Category']
max = ws1.max_row


if Month != (previous_execution_month.value):
    ws1.merge_cells(start_row=max+1, start_column=1, end_row=max+2, end_column=1)
    ws1.cell(row=max+1, column=1).value = year_month_format

    ws1.cell(row=max+1, column=2).value = 'EDW'
    ws1.cell(row=max+2, column=2).value = 'AWS'

    currentCell = ws1.cell(row=max+1, column=2)
    currentCell.alignment = Alignment(horizontal='center',vertical='center')
    currentCell = ws1.cell(row=max+2, column=2)
    currentCell.alignment = Alignment(horizontal='center',vertical='center')
    currentCell = ws1.cell(row=max+1, column=1)
    currentCell.alignment = Alignment(horizontal='center',vertical='center')

    ws1.cell(row=max+1, column=3).value = DQ_Result[0][0]
    ws1.cell(row=max+1, column=4).value = DQ_Result[0][1]
    ws1.cell(row=max+1, column=5).value = DQ_Result[0][2]
    ws1.cell(row=max+1, column=6).value = DQ_Result[0][3]
    ws1.cell(row=max+1, column=7).value = DQ_Result[0][4]
    ws1.merge_cells(start_row=max+1, start_column=8, end_row=max+2, end_column=8)
    ws1.merge_cells(start_row=max+1, start_column=9, end_row=max+2, end_column=9)
    ws1.merge_cells(start_row=max+1, start_column=10, end_row=max+2, end_column=10)
    ws1.cell(row=max+1, column=8).value = DQ_Result[0][5]
    ws1.cell(row=max+1, column=9).value = DQ_Result[0][6]
    ws1.cell(row=max+1, column=10).value = DQ_Result[0][7]
    ws1.cell(row=max+2, column=3).value = DQ_Result[1][0]
    ws1.cell(row=max+2, column=4).value = DQ_Result[1][1]
    ws1.cell(row=max+2, column=5).value = DQ_Result[1][2]
    ws1.cell(row=max+2, column=6).value = DQ_Result[1][3]
    ws1.cell(row=max+2, column=7).value = DQ_Result[1][4]
    currentCell = ws1.cell(row=max+1, column=8) #or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center',vertical='center')
    currentCell = ws1.cell(row=max+1, column=9) #or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center',vertical='center')
    currentCell = ws1.cell(row=max+1, column=10) #or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center',vertical='center')

    ws1 = wb['Label=None']
    ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=7)
    ws1.cell(row=1, column=1).value = "List of Tickets with Label = None"
    currentCell = ws1.cell(row=1, column=1)
    currentCell.alignment = Alignment(horizontal='center',vertical='center')
    currentCell.fill = PatternFill(fill_type='solid', start_color='FF6347', end_color='FF6347')
    max = ws1.max_row
    ws1.cell(row=max+1, column=1).value = year_month_format
    ws1.cell(row=max+1, column=2).value = 'None'
    ws1.cell(row=max+1, column=3).value = DQ_Null_Report[0]
    ws1.cell(row=max+1, column=4).value = DQ_Null_Report[1]
    ws1.cell(row=max+1, column=5).value = DQ_Null_Report[2]
    ws1.cell(row=max+1, column=6).value = DQ_Null_Report[3]
    ws1.cell(row=max+1, column=7).value = total_none

else:
    max = max - 1
    print("ELSE MAX", max)
    ws1.cell(row=max, column=1).value = year_month_format

    ws1.cell(row=max, column=2).value = 'EDW'
    ws1.cell(row=max + 1, column=2).value = 'AWS'

    currentCell = ws1.cell(row=max, column=2)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    currentCell = ws1.cell(row=max + 1, column=2)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    currentCell = ws1.cell(row=max, column=1)
    currentCell.alignment = Alignment(horizontal='center', vertical='center')

    ws1.cell(row=max, column=3).value = DQ_Result[0][0]
    ws1.cell(row=max, column=4).value = DQ_Result[0][1]
    ws1.cell(row=max, column=5).value = DQ_Result[0][2]
    ws1.cell(row=max, column=6).value = DQ_Result[0][3]
    ws1.cell(row=max, column=7).value = DQ_Result[0][4]
    ws1.cell(row=max, column=8).value = DQ_Result[0][5]
    ws1.cell(row=max, column=9).value = DQ_Result[0][6]
    ws1.cell(row=max, column=10).value = DQ_Result[0][7]
    ws1.cell(row=max + 1, column=3).value = DQ_Result[1][0]
    ws1.cell(row=max + 1, column=4).value = DQ_Result[1][1]
    ws1.cell(row=max + 1, column=5).value = DQ_Result[1][2]
    ws1.cell(row=max + 1, column=6).value = DQ_Result[1][3]
    ws1.cell(row=max + 1, column=7).value = DQ_Result[1][4]
    currentCell = ws1.cell(row=max , column=8)  # or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    currentCell = ws1.cell(row=max , column=9)  # or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center', vertical='center')
    currentCell = ws1.cell(row=max , column=10)  # or currentCell = ws['A1']
    currentCell.alignment = Alignment(horizontal='center', vertical='center')

    ws1 = wb['Label=None']
    max = ws1.max_row
    ws1.cell(row=max, column=1).value = year_month_format
    ws1.cell(row=max, column=2).value = 'None'
    ws1.cell(row=max, column=3).value = DQ_Null_Report[0]
    ws1.cell(row=max, column=4).value = DQ_Null_Report[1]
    ws1.cell(row=max, column=5).value = DQ_Null_Report[2]
    ws1.cell(row=max, column=6).value = DQ_Null_Report[3]
    ws1.cell(row=max, column=7).value = total_none

wb.save(filename=workbook_name)
print("Completed Writing Issues to the Excel File - Ticket_Counts")

print("Updating Excel - current Month value")
ws1 = wb['List of tickets']
c3 = ws1['O2']
c3.value = Month
wb.save(filename=workbook_name)

def send_mail(send_from,send_to,subject,text,server,port,username='',password='',isTls=True):
    print("Mail Part Testing")
    # try:
    #     msg = MIMEMultipart()
    #     msg['From'] = send_from
    #     msg['To'] = send_to
    #     msg['Date'] = formatdate(localtime = True)
    #     msg['Subject'] = subject
    #     msg.attach(MIMEText(text))
    #
    #     part = MIMEBase('application', "octet-stream")
    #     part.set_payload(open("Ticket_Counts.xlsx", "rb").read())
    #     encoders.encode_base64(part)
    #     part.add_header('Content-Disposition', 'attachment; filename="Ticket_Counts.xlsx"')
    #     msg.attach(part)
    #
    #     #context = ssl.SSLContext(ssl.PROTOCOL_SSLv3)
    #     #SSL connection only working on Python 3+
    #     smtp = smtplib.SMTP(server, port)
    #     if isTls:
    #         smtp.starttls()
    #     smtp.login(username,password)
    #     smtp.sendmail(send_from, send_to, msg.as_string())
    #     smtp.quit()
    #     print("Email Sent")
    #     completed = True
    # except Exception as e:
    #     print("Could not send the email. Please Check!")
    #     print(e)
    #     if completed:
    #         print("Script Executed Successfully")
    #         pass
    #     else:
    #         workbook_name = 'Ticket_Counts.xlsx'
    #         wb = load_workbook(workbook_name)
    #         ws1 = wb['List of tickets']
    #         c3 = ws1['O2']
    #         c3.value = Month - 1
    #         wb.save(filename=workbook_name)



def main():
    send_mail(config['from'],config['to'],config['subject_keyword_DQ_Quality_Tickets'], "Additional Information to be added", config['server'], config['port'], config['login'], config['password'])

if __name__ == '__main__':
    main()